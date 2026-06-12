# inventory_forecast_dss_app.py
# Full DSS – EDA • ABC–XYZ • Forecast • Inventory • Time Series • Summary Dashboard
# Requires: streamlit, pandas, numpy, matplotlib, statsmodels, xgboost, catboost, openpyxl, scipy

import io
import math
import warnings
import itertools
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

import streamlit as st
from scipy import stats

from statsmodels.tsa.statespace.sarimax import SARIMAX
from xgboost import XGBRegressor
from catboost import CatBoostRegressor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

# =============================================================================
# GLOBAL CONFIG
# =============================================================================

# Inventory / EOQ data file (must be in working directory)
EXCEL_FILE_EOQ = "ForecastEOQ.xlsx"
OUTPUT_DIR = "./"

MATERIAL_CODES = [600080, 600096, 600102, 600112, 603789, 603812, 607290, 626100, 627518, 627519]

SHORT_NAMES: Dict[int, str] = {
    600080: "MOTORIN",
    600096: "TOKA 32",
    600102: "ÇEMBER TOKA",
    600112: "SIYAH ÇEMBER",
    603789: "ELDİVEN",
    603812: "KULAK TIKACI",
    607290: "Ç.4140",
    626100: "SAPAN",
    627518: "BRK 77x60",
    627519: "BRK 80x122",
}

# EOQ parameters
ORDERING_COST = 2792          # TRY per order
HOLDING_COST_PCT = 0.25       # 25% annual
SERVICE_LEVEL = 0.95          # 95%
STOCKOUT_COST_PER_UNIT = 500  # TRY per unit stockout

COLORS = [
    "#2563EB", "#DC2626", "#16A34A", "#D97706", "#7C3AED",
    "#0891B2", "#DB2777", "#65A30D", "#EA580C", "#6366F1",
]

# Forecast (Motorin) configuration
DATE_COL = "Tarih"
VALUE_COL = "Tüketim"
TEST_START = "2025-04-02"  # first predicted day
SEASONAL = 7
DEFAULT_ORDER = (1, 1, 2)
DEFAULT_SORDER = (1, 1, 1, SEASONAL)
ALPHA = 0.05
AUTO_SELECT = False

N_LAGS = 14
ROLL_WINDOWS = (7, 14, 30)


# =============================================================================
# INVENTORY / EOQ FUNCTIONS
# =============================================================================

@st.cache_data
def load_material_data(filepath: str, material_code: int) -> Tuple[pd.DataFrame, Dict]:
    """
    Load real data for a material from ForecastEOQ.xlsx.
    Columns: Tarih, Mal Giriş, Tüketim Miktarı, Unit Price, Para Birimi, Lead time
    """
    try:
        df = pd.read_excel(filepath, sheet_name=str(material_code))
        df.columns = df.columns.str.strip()
        df["Tarih"] = pd.to_datetime(df["Tarih"])
        df = df.sort_values("Tarih").reset_index(drop=True)

        df["Mal Giriş"] = df["Mal Giriş"].fillna(0)
        df["Tüketim Miktarı"] = df["Tüketim Miktarı"].fillna(0)

        unit_price = float(df["Unit Price"].iloc[0]) if "Unit Price" in df.columns else 1.0
        currency = str(df["Para Birimi"].iloc[0]) if "Para Birimi" in df.columns else "TRY"
        lead_time = int(df["Lead time"].iloc[0]) if "Lead time" in df.columns else 2

        metadata = {
            "unit_price": unit_price,
            "currency": currency,
            "lead_time": lead_time,
            "material_code": material_code,
            "material_name": SHORT_NAMES.get(material_code, str(material_code)),
        }
        return df, metadata
    except Exception as e:
        st.error(f"Error reading material code {material_code}: {e}")
        return None, None


def calculate_real_inventory(df: pd.DataFrame) -> pd.DataFrame:
    """
    Real Inventory = Previous Inventory + Inflow - Consumption
    Allows negative inventory (stockout) to show below zero line.
    """
    df = df.copy()
    df["Envanter"] = 0.0
    df["Stockout"] = 0.0

    current_inv = 0.0
    inventory_list = []
    stockout_list = []

    for _, row in df.iterrows():
        current_inv = current_inv + row["Mal Giriş"] - row["Tüketim Miktarı"]
        if current_inv < 0:
            stockout_list.append(-current_inv)
        else:
            stockout_list.append(0)
        inventory_list.append(current_inv)

    df["Envanter"] = inventory_list
    df["Stockout"] = stockout_list
    return df


def calculate_real_inventory_costs(df: pd.DataFrame, metadata: Dict) -> Tuple[pd.DataFrame, Dict]:
    """
    Real inventory costs:
    - Holding (positive inventory)
    - Ordering (inflow > 0)
    - Stockout (shortage units)
    """
    df = df.copy()
    unit_price = metadata["unit_price"]

    daily_holding_cost_rate = HOLDING_COST_PCT / 365
    df["Holding_Cost_Daily"] = df["Envanter"].clip(lower=0) * daily_holding_cost_rate * unit_price

    df["Is_Order"] = df["Mal Giriş"] > 0
    df["Ordering_Cost_Daily"] = df["Is_Order"] * ORDERING_COST

    df["Stockout_Cost_Daily"] = df["Stockout"] * STOCKOUT_COST_PER_UNIT

    df["Total_Cost_Daily"] = df["Holding_Cost_Daily"] + df["Ordering_Cost_Daily"] + df["Stockout_Cost_Daily"]

    total_holding_cost = df["Holding_Cost_Daily"].sum()
    total_ordering_cost = df["Ordering_Cost_Daily"].sum()
    total_stockout_cost = df["Stockout_Cost_Daily"].sum()
    total_cost = total_holding_cost + total_ordering_cost + total_stockout_cost

    days = len(df)
    avg_daily_cost = total_cost / days if days > 0 else 0
    annual_cost = avg_daily_cost * 365

    total_orders = df["Is_Order"].sum()
    total_stockout_units = df["Stockout"].sum()
    stockout_days = (df["Stockout"] > 0).sum()

    cost_summary = {
        "material_code": metadata["material_code"],
        "material_name": metadata["material_name"],
        "days_analyzed": days,
        "total_holding_cost": round(total_holding_cost, 2),
        "total_ordering_cost": round(total_ordering_cost, 2),
        "total_stockout_cost": round(total_stockout_cost, 2),
        "total_cost": round(total_cost, 2),
        "avg_daily_cost": round(avg_daily_cost, 2),
        "annual_cost": round(annual_cost, 2),
        "total_orders": int(total_orders),
        "avg_order_qty": round(df[df["Mal Giriş"] > 0]["Mal Giriş"].mean(), 2) if total_orders > 0 else 0,
        "total_stockout_units": round(total_stockout_units, 2),
        "stockout_days": int(stockout_days),
        "unit_price": metadata["unit_price"],
        "currency": metadata["currency"],
    }
    return df, cost_summary


def calculate_statistics(df: pd.DataFrame, metadata: Dict) -> Dict:
    """
    EOQ & ROP based on real data.
    """
    consumption = df["Tüketim Miktarı"].values
    inflow = df["Mal Giriş"].values
    inventory = df["Envanter"].values

    days = len(df)
    total_consumption = consumption.sum()
    total_inflow = inflow.sum()
    avg_daily_consumption = total_consumption / days if days > 0 else 0

    consumption_nonzero = consumption[consumption > 0]
    std_consumption = consumption_nonzero.std() if len(consumption_nonzero) > 1 else avg_daily_consumption * 0.3
    std_daily = std_consumption / np.sqrt(len(df)) if len(df) > 1 else avg_daily_consumption * 0.3

    cv = (std_daily / avg_daily_consumption * 100) if avg_daily_consumption > 0 else 0

    annual_demand = avg_daily_consumption * 365
    holding_cost = HOLDING_COST_PCT * metadata["unit_price"]

    if annual_demand > 0 and holding_cost > 0:
        eoq = math.sqrt((2 * annual_demand * ORDERING_COST) / holding_cost)
    else:
        eoq = avg_daily_consumption * 30
    eoq = max(1.0, eoq)

    lead_time = metadata["lead_time"]
    lt_mean = avg_daily_consumption * lead_time
    lt_std = std_daily * math.sqrt(lead_time)

    z = stats.norm.ppf(SERVICE_LEVEL)
    rop = max(0.0, lt_mean + z * lt_std)

    return {
        "material_code": metadata["material_code"],
        "material_name": metadata["material_name"],
        "total_days": days,
        "total_consumption": total_consumption,
        "total_inflow": total_inflow,
        "avg_daily_consumption": round(avg_daily_consumption, 2),
        "std_daily": round(std_daily, 2),
        "cv_percent": round(cv, 1),
        "max_inventory": float(inventory.max()),
        "min_inventory": float(inventory.min()),
        "avg_inventory": float(inventory.mean()),
        "eoq": round(eoq, 2),
        "rop": round(rop, 2),
        "unit_price": metadata["unit_price"],
        "currency": metadata["currency"],
        "lead_time": lead_time,
        "annual_demand": round(annual_demand, 0),
    }


def export_inventory_cost_data(df: pd.DataFrame, cost_summary: Dict) -> bytes:
    """
    Export detailed inventory cost data to Excel (as bytes for download).
    """
    export_df = pd.DataFrame({
        "Date": df["Tarih"],
        "Inflow": df["Mal Giriş"],
        "Consumption": df["Tüketim Miktarı"],
        "Inventory": df["Envanter"],
        "Stockout": df["Stockout"],
        "Holding_Cost_Daily": df["Holding_Cost_Daily"].round(2),
        "Ordering_Cost_Daily": df["Ordering_Cost_Daily"].round(2),
        "Stockout_Cost_Daily": df["Stockout_Cost_Daily"].round(2),
        "Total_Cost_Daily": df["Total_Cost_Daily"].round(2),
    })

    material_name = cost_summary["material_name"].replace(" ", "_")
    material_code = cost_summary["material_code"]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name=f"{material_code}_{material_name}", index=False)
    buffer.seek(0)
    return buffer.read()


def create_eoq_graph(df: pd.DataFrame, stats_d: Dict, cost_summary: Dict):
    """
    EOQ saw-tooth graph with negative inventory zone.
    """
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(16, 10))
    fig.patch.set_facecolor("#F8FAFC")

    ax1.set_facecolor("#FFFFFF")
    days = np.arange(len(df))

    ax1.fill_between(days, 0, stats_d["rop"], alpha=0.08, color="#16A34A",
                     label="Safety Stock Zone")

    ax1.fill_between(days, 0, df["Envanter"], where=(df["Envanter"] < 0),
                     alpha=0.15, color="#EF4444", label="Stockout Zone")

    ax1.plot(days, df["Envanter"], color="#DC2626", linewidth=2.0,
             label="Real Inventory Level", marker="o", markersize=2,
             alpha=0.85, zorder=3)

    ax1.axhline(y=stats_d["rop"], color="#16A34A", linestyle="--", linewidth=2.0,
                label=f"ROP = {stats_d['rop']:.0f}", alpha=0.8)
    ax1.axhline(y=stats_d["eoq"], color="#D97706", linestyle="--", linewidth=2.0,
                label=f"EOQ = {stats_d['eoq']:.0f}", alpha=0.7)
    ax1.axhline(y=stats_d["avg_inventory"], color="#7C3AED", linestyle=":",
                linewidth=2.0, label=f"Avg Inv = {stats_d['avg_inventory']:.0f}", alpha=0.8)
    ax1.axhline(y=0, color="#000000", linestyle="-", linewidth=1.2, alpha=0.5)

    ax1.set_xlabel("Days", fontsize=12, fontweight="bold", color="#374151")
    ax1.set_ylabel("Inventory (units)", fontsize=12, fontweight="bold", color="#374151")
    ax1.set_title(
        f"📦 EOQ Saw-Tooth Graph — {stats_d['material_name']}",
        fontsize=13, fontweight="bold", color="#111827", pad=10
    )
    ax1.legend(fontsize=9, loc="upper right", framealpha=0.95)
    ax1.grid(True, alpha=0.25, linestyle="--", color="#9CA3AF")
    ax1.spines[["top", "right"]].set_visible(False)

    info_text = (
        f"Code: {stats_d['material_code']}\n"
        f"Lead Time: {stats_d['lead_time']} days\n"
        f"Unit Price: {stats_d['currency']}{stats_d['unit_price']:.2f}\n"
        f"Avg Daily Cons.: {stats_d['avg_daily_consumption']:.2f}\n"
        f"Std Dev: {stats_d['std_daily']:.2f}\n"
        f"CV: {stats_d['cv_percent']:.1f}%\n"
        f"Service Level: {SERVICE_LEVEL*100:.0f}%\n"
        f"Total Cost: {stats_d['currency']}{cost_summary['total_cost']:,.2f}\n"
        f"Stockout Days: {cost_summary['stockout_days']}"
    )

    ax1.text(
        0.02, 0.98, info_text, transform=ax1.transAxes,
        fontsize=9, ha="left", va="top", fontweight="bold", color="#111827",
        bbox=dict(boxstyle="round,pad=0.8", facecolor="#FEF3C7", alpha=0.95,
                  edgecolor="#D97706", linewidth=2)
    )

    ax2.set_facecolor("#FFFFFF")
    x = np.arange(len(df))
    width = 0.35

    ax2.bar(x - width/2, df["Mal Giriş"], width, label="Inflow",
            color="#16A34A", alpha=0.8, edgecolor="white")
    ax2.bar(x + width/2, df["Tüketim Miktarı"], width, label="Consumption",
            color="#DC2626", alpha=0.8, edgecolor="white")

    ax2.set_xlabel("Days", fontsize=12, fontweight="bold", color="#374151")
    ax2.set_ylabel("Quantity (units)", fontsize=12, fontweight="bold", color="#374151")
    ax2.set_title("Daily Inflow vs Consumption", fontsize=12, fontweight="bold",
                  color="#111827", pad=10)
    ax2.legend(fontsize=9, loc="upper right")
    ax2.grid(True, alpha=0.2, linestyle="--", axis="y")
    ax2.spines[["top", "right"]].set_visible(False)

    if len(df) > 50:
        step = len(df) // 10
        ax2.set_xticks(range(0, len(df), step))

    plt.tight_layout()
    return fig


def create_eoq_cost_graph(stats_d: Dict, cost_summary: Dict):
    """
    EOQ cost tradeoff + real cost comparison.
    """
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
    fig.patch.set_facecolor("#F8FAFC")
    ax1.set_facecolor("#FFFFFF")

    annual_demand = stats_d["annual_demand"]
    eoq_value = stats_d["eoq"]
    unit_price = stats_d["unit_price"]

    Q_range = np.linspace(1, eoq_value * 3, 300)
    ordering_cost = (annual_demand / Q_range) * ORDERING_COST
    holding_cost = (Q_range / 2) * HOLDING_COST_PCT * unit_price
    total_cost = ordering_cost + holding_cost

    ax1.plot(Q_range, ordering_cost, linewidth=2.0, color="#DC2626",
             label="Ordering Cost", linestyle="--", alpha=0.8)
    ax1.plot(Q_range, holding_cost, linewidth=2.0, color="#16A34A",
             label="Holding Cost", linestyle="--", alpha=0.8)
    ax1.plot(Q_range, total_cost, linewidth=3.0, color="#7C3AED",
             label="Total Cost", zorder=3)

    eoq_cost = (eoq_value / 2) * HOLDING_COST_PCT * unit_price + \
               (annual_demand / eoq_value) * ORDERING_COST

    ax1.scatter([eoq_value], [eoq_cost], s=250, color="#D97706", marker="*",
                zorder=4, edgecolor="black", linewidth=1.5,
                label=f"EOQ = {eoq_value:.0f}")

    ax1.axvline(x=eoq_value, color="#D97706", linestyle=":", linewidth=2, alpha=0.7)

    ax1.set_xlabel("Order Quantity (Q)", fontsize=11, fontweight="bold", color="#374151")
    ax1.set_ylabel(f"Annual Cost ({stats_d['currency']})", fontsize=11, fontweight="bold", color="#374151")
    ax1.set_title("EOQ Cost Tradeoff Function", fontsize=12, fontweight="bold", color="#111827")
    ax1.legend(fontsize=9, loc="upper right", framealpha=0.95)
    ax1.grid(True, alpha=0.2, linestyle="--")
    ax1.spines[["top", "right"]].set_visible(False)

    ax2.axis("off")

    annual_ordering = (annual_demand / eoq_value) * ORDERING_COST
    annual_holding = (eoq_value / 2) * HOLDING_COST_PCT * unit_price
    total_annual = annual_ordering + annual_holding

    text = f"""
EOQ COST ANALYSIS — {stats_d['material_name']}

Demand & Parameters
-------------------
Annual Demand:           {annual_demand:>12,.0f} units
Avg Daily Consumption:   {stats_d['avg_daily_consumption']:>12,.2f} units
Lead Time:               {stats_d['lead_time']:>12} days
Unit Price:              {stats_d['currency']}{unit_price:>12,.2f}
CV:                      {stats_d['cv_percent']:>12.1f}%

EOQ (Optimal Order Quantity)
----------------------------
Ordering Cost per Order: {stats_d['currency']}{ORDERING_COST:>12,.2f}
Holding Cost Rate:       {HOLDING_COST_PCT*100:>12.1f}%

EOQ Formula: Q* = √(2DS/h)
  D = {annual_demand:,.0f}
  S = {stats_d['currency']}{ORDERING_COST:,.2f}
  h = {stats_d['currency']}{HOLDING_COST_PCT * unit_price:,.3f}

Optimal Order Quantity:  {eoq_value:>12,.0f} units

Annual Cost (Theoretical)
-------------------------
Ordering Cost:           {stats_d['currency']}{annual_ordering:>12,.2f}
Holding Cost:            {stats_d['currency']}{annual_holding:>12,.2f}
TOTAL ANNUAL COST:       {stats_d['currency']}{total_annual:>12,.2f}

Annual Cost (Real Data)
-----------------------
Holding Cost:            {stats_d['currency']}{cost_summary['total_holding_cost'] * (365/cost_summary['days_analyzed']):>12,.2f}
Ordering Cost:           {stats_d['currency']}{cost_summary['total_ordering_cost'] * (365/cost_summary['days_analyzed']):>12,.2f}
Stockout Cost:           {stats_d['currency']}{cost_summary['total_stockout_cost'] * (365/cost_summary['days_analyzed']):>12,.2f}
TOTAL ANNUAL COST:       {stats_d['currency']}{cost_summary['annual_cost']:>12,.2f}

Average Inventory:       {eoq_value/2:>12,.0f} units
"""

    ax2.text(
        0.02, 0.98, text, transform=ax2.transAxes,
        fontsize=8.5, ha="left", va="top", family="monospace", color="#111827",
        fontweight="bold",
        bbox=dict(boxstyle="round,pad=1", facecolor="#FEF3C7", alpha=0.95,
                  edgecolor="#D97706", linewidth=2)
    )

    plt.tight_layout()
    return fig


def create_summary_table(all_stats: List[Dict], all_costs: List[Dict]) -> pd.DataFrame:
    df_stats = pd.DataFrame(all_stats)
    df_costs = pd.DataFrame(all_costs)
    df_summary = df_stats.merge(
        df_costs[[
            "material_code", "total_holding_cost", "total_ordering_cost",
            "total_stockout_cost", "total_cost", "avg_daily_cost", "annual_cost",
            "total_orders", "total_stockout_units", "stockout_days"
        ]],
        on="material_code"
    )
    df_summary = df_summary.sort_values("annual_demand", ascending=False)
    return df_summary


# =============================================================================
# FORECAST (MOTORIN) FUNCTIONS – from forecast_eren_11_06
# =============================================================================

def load_daily_from_bytes(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))
    df[DATE_COL] = pd.to_datetime(df[DATE_COL])
    daily = (
        df.set_index(DATE_COL)[VALUE_COL]
        .sort_index()
        .asfreq("D")
        .fillna(0)
    )
    out = pd.DataFrame({"date": daily.index, VALUE_COL: daily.values})
    out["t"] = np.arange(1, len(out) + 1)
    return out


def split_train_test(frame: pd.DataFrame):
    cutoff = pd.Timestamp(TEST_START)
    train = frame[frame["date"] < cutoff].reset_index(drop=True)
    test = frame[frame["date"] >= cutoff].reset_index(drop=True)
    return train, test


def _aicc(result) -> float:
    k = len(result.params)
    n = result.nobs
    return result.aic + (2 * k * k + 2 * k) / (n - k - 1) if (n - k - 1) > 0 else np.inf


def select_order_aicc(y: pd.Series):
    best = {"score": np.inf, "order": DEFAULT_ORDER, "sorder": DEFAULT_SORDER}
    for p, d, q in itertools.product(range(3), range(2), range(3)):
        for P, D, Q in itertools.product(range(2), range(2), range(2)):
            try:
                res = SARIMAX(
                    y, order=(p, d, q),
                    seasonal_order=(P, D, Q, SEASONAL),
                    enforce_stationarity=False,
                    enforce_invertibility=False
                ).fit(disp=False)
                score = _aicc(res)
                if np.isfinite(score) and score < best["score"]:
                    best = {"score": score, "order": (p, d, q),
                            "sorder": (P, D, Q, SEASONAL)}
            except Exception:
                continue
    return best["order"], best["sorder"]


def rolling_one_step(y_train: pd.Series, test_values: np.ndarray, order, sorder):
    res = SARIMAX(
        y_train, order=order, seasonal_order=sorder,
        enforce_stationarity=False, enforce_invertibility=False
    ).fit(disp=False)

    hist = res
    pred = np.empty(len(test_values))
    lo = np.empty(len(test_values))
    hi = np.empty(len(test_values))

    for i, actual in enumerate(test_values):
        fc = hist.get_forecast(steps=1)
        pred[i] = float(fc.predicted_mean.iloc[0])
        ci = np.asarray(fc.conf_int(alpha=ALPHA))[0]
        lo[i], hi[i] = ci[0], ci[1]
        hist = hist.append([actual], refit=False)

    lo = np.maximum(lo, 0.0)
    hi = np.maximum(hi, 0.0)
    return pred, lo, hi


def make_feature_row(history_values, idx_date, t_value):
    h = np.asarray(history_values, dtype=float)
    feat = {}
    for lag in range(1, N_LAGS + 1):
        feat[f"lag_{lag}"] = h[-lag] if len(h) >= lag else 0.0
    for w in ROLL_WINDOWS:
        window = h[-w:] if len(h) >= 1 else np.array([0.0])
        feat[f"rmean_{w}"] = float(np.mean(window)) if len(window) else 0.0
        feat[f"rstd_{w}"] = float(np.std(window)) if len(window) > 1 else 0.0

    dow = idx_date.dayofweek
    feat["dow"] = dow
    feat["dow_sin"] = np.sin(2 * np.pi * dow / 7)
    feat["dow_cos"] = np.cos(2 * np.pi * dow / 7)
    feat["month"] = idx_date.month
    feat["month_sin"] = np.sin(2 * np.pi * idx_date.month / 12)
    feat["month_cos"] = np.cos(2 * np.pi * idx_date.month / 12)
    feat["trend"] = t_value
    return feat


def build_training_matrix(train_frame):
    values = train_frame[VALUE_COL].values.astype(float)
    dates = train_frame["date"].values
    ts = train_frame["t"].values

    rows, targets = [], []
    for i in range(N_LAGS, len(values)):
        hist = values[:i]
        feat = make_feature_row(hist, pd.Timestamp(dates[i]), int(ts[i]))
        rows.append(feat)
        targets.append(values[i])

    X = pd.DataFrame(rows)
    y = np.asarray(targets, dtype=float)
    return X, y, list(X.columns)


def rolling_one_step_ml(model, train_frame, test_frame, feature_cols):
    history = list(train_frame[VALUE_COL].values.astype(float))
    preds = []
    for _, row in test_frame.iterrows():
        feat = make_feature_row(history, pd.Timestamp(row["date"]), int(row["t"]))
        x = pd.DataFrame([feat])[feature_cols]
        yhat = max(0.0, float(model.predict(x)[0]))
        preds.append(yhat)
        history.append(float(row[VALUE_COL]))
    return np.asarray(preds)


def fit_xgboost(X, y):
    model = XGBRegressor(
        n_estimators=600, learning_rate=0.03, max_depth=4,
        subsample=0.8, colsample_bytree=0.8, min_child_weight=3,
        reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbosity=0
    )
    model.fit(X, y, verbose=False)
    return model


def fit_catboost(X, y):
    model = CatBoostRegressor(
        iterations=600, learning_rate=0.03, depth=4,
        l2_leaf_reg=3, subsample=0.8, random_seed=42, verbose=0
    )
    model.fit(X, y)
    return model


def error_stats(actual, pred):
    actual = np.asarray(actual, dtype=float)
    pred = np.asarray(pred, dtype=float)
    err = actual - pred
    mae = float(np.mean(np.abs(err)))
    rmse = float(np.sqrt(np.mean(err ** 2)))
    mask = actual != 0
    mape = float(np.mean(np.abs(err[mask] / actual[mask])) * 100) if mask.any() else np.nan
    return mae, rmse, mape


def build_comparison(test, sarima_pred, lo, hi, xgb_pred, cat_pred) -> pd.DataFrame:
    out = pd.DataFrame({
        "date": test["date"].values,
        "t": test["t"].values,
        "Actual": test[VALUE_COL].values.astype(float),
        "SARIMA": sarima_pred,
        "SARIMA_Lo95": lo,
        "SARIMA_Hi95": hi,
        "XGBoost": xgb_pred,
        "CatBoost": cat_pred,
    })
    return out


def plot_results(frame, train, comp, order, sorder, stats):
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(13, 9))
    boundary = train["date"].iloc[-1]

    ax1.plot(frame["date"], frame[VALUE_COL], color="#1f77b4",
             linewidth=0.7, alpha=0.55, label="Actual (daily)")
    ax1.plot(comp["date"], comp["SARIMA"], color="#2ca02c",
             linewidth=1.3, label="SARIMA")
    ax1.plot(comp["date"], comp["XGBoost"], color="#d62728",
             linewidth=1.3, linestyle="--", label="XGBoost")
    ax1.plot(comp["date"], comp["CatBoost"], color="#9467bd",
             linewidth=1.3, linestyle="-.", label="CatBoost")

    ax1.axvline(boundary, color="grey", linestyle=":", linewidth=1.3)
    ax1.text(boundary, ax1.get_ylim()[1], " train | test",
             color="grey", va="top", ha="left", fontsize=9)

    ax1.set_title("Baseline Historical Time Series Data — Three Models",
                  fontsize=11, fontweight="bold")
    ax1.set_ylabel("Daily consumption")
    ax1.legend(loc="upper left", framealpha=0.9, fontsize=9)
    ax1.grid(True, alpha=0.3)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
    ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=2))

    ax2.fill_between(comp["date"], comp["SARIMA_Lo95"], comp["SARIMA_Hi95"],
                     color="#2ca02c", alpha=0.10, label="SARIMA 95% interval")
    ax2.plot(comp["date"], comp["Actual"], color="#1f77b4",
             linewidth=1.0, marker="o", markersize=2.5, label="Actual (daily)")
    ax2.plot(comp["date"], comp["SARIMA"], color="#2ca02c",
             linewidth=1.6, label="SARIMA")
    ax2.plot(comp["date"], comp["XGBoost"], color="#d62728",
             linewidth=1.6, linestyle="--", label="XGBoost")
    ax2.plot(comp["date"], comp["CatBoost"], color="#9467bd",
             linewidth=1.6, linestyle="-.", label="CatBoost")

    pstart = pd.Timestamp(comp["date"].iloc[0]).strftime("%Y-%m-%d")
    ax2.set_title(
        f"Prediction window from {pstart}: real vs rolling predictions (daily)",
        fontsize=11, fontweight="bold"
    )
    ax2.set_xlabel("Date")
    ax2.set_ylabel("Daily consumption")
    ax2.legend(loc="upper left", framealpha=0.9, fontsize=9)
    ax2.grid(True, alpha=0.3)
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    ax2.xaxis.set_major_locator(mdates.AutoDateLocator())

    sub = (
        f"SARIMA MAE={stats['SARIMA'][0]:,.0f} RMSE={stats['SARIMA'][1]:,.0f} | "
        f"XGBoost MAE={stats['XGBoost'][0]:,.0f} RMSE={stats['XGBoost'][1]:,.0f} | "
        f"CatBoost MAE={stats['CatBoost'][0]:,.0f} RMSE={stats['CatBoost'][1]:,.0f}"
    )

    fig.suptitle(
        f"Daily Motorin Consumption — SARIMA{order}{sorder[:3]} s={sorder[3]} "
        f"vs XGBoost vs CatBoost (rolling 1-step)\n{sub}",
        fontsize=12, fontweight="bold"
    )
    fig.autofmt_xdate()
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    return fig


def write_excel_forecast(comp, order, sorder, stats) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Comparison"

    arial = "Arial"
    body = Font(name=arial)
    head_font = Font(name=arial, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Rolling 1-step Forecast — Daily Motorin: SARIMA vs XGBoost vs CatBoost"
    ws["A1"].font = Font(name=arial, bold=True, size=13)
    ws["A2"] = f"SARIMA{order}{sorder[:3]} s={sorder[3]} (rows as t = 1, 2, 3, … , N)"
    ws["A2"].font = Font(name=arial, italic=True, color="595959")

    headers = ["Date", "t", "Actual", "SARIMA", "XGBoost", "CatBoost",
               "SARIMA Lo95", "SARIMA Hi95"]
    hr = 4
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(hr, c, name)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border

    first = hr + 1
    for i, row in comp.iterrows():
        r = first + i
        ws.cell(r, 1, pd.Timestamp(row["date"]).strftime("%Y-%m-%d")).font = body
        ws.cell(r, 2, int(row["t"])).font = body
        ws.cell(r, 3, round(float(row["Actual"]))).font = body
        ws.cell(r, 4, round(float(row["SARIMA"]))).font = body
        ws.cell(r, 5, round(float(row["XGBoost"]))).font = body
        ws.cell(r, 6, round(float(row["CatBoost"]))).font = body
        ws.cell(r, 7, round(float(row["SARIMA_Lo95"]))).font = body
        ws.cell(r, 8, round(float(row["SARIMA_Hi95"]))).font = body
        for c in range(1, 9):
            ws.cell(r, c).border = border
        for c in (3, 4, 5, 6, 7, 8):
            ws.cell(r, c).number_format = "#,##0"

    last = first + len(comp) - 1
    summary = last + 2
    ws.cell(summary, 1, "Model Accuracy (test set, rolling 1-step)").font = Font(name=arial, bold=True, size=12)
    th = summary + 1
    for c, name in enumerate(["Model", "MAE", "RMSE", "MAPE (%)"], start=1):
        cell = ws.cell(th, c, name)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border

    for j, model_name in enumerate(["SARIMA", "XGBoost", "CatBoost"]):
        r = th + 1 + j
        mae, rmse, mape = stats[model_name]
        ws.cell(r, 1, model_name).font = Font(name=arial, bold=True)
        ws.cell(r, 2, round(mae)).font = body
        ws.cell(r, 3, round(rmse)).font = body
        ws.cell(r, 4, round(mape, 1) if np.isfinite(mape) else "").font = body
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "#,##0"
        ws.cell(r, 4).number_format = "0.0"
        for c in range(1, 5):
            ws.cell(r, c).border = border

    for c, w in enumerate([12, 6, 11, 11, 11, 11, 12, 12], start=1):
        ws.column_dimensions[chr(64 + c)].width = w

    ws.row_dimensions[hr].height = 28
    ws.freeze_panes = "A5"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# =============================================================================
# STREAMLIT APP
# =============================================================================

def main():
    st.set_page_config(
        page_title="Inventory & Forecast DSS",
        layout="wide",
        page_icon="📦",
    )

    st.markdown(
        "<h1 style='text-align:center;'>📦 Integrated Inventory & Forecast Decision Support System</h1>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p style='text-align:center;color:#6b7280;'>EDA • ABC–XYZ • Forecast • Inventory • Time Series • Summary Dashboard</p>",
        unsafe_allow_html=True,
    )

    # Sidebar – material selection for inventory-related views
    st.sidebar.header("Material Selector")
    material_code = st.sidebar.selectbox(
        "Target Material Code",
        options=MATERIAL_CODES,
        format_func=lambda x: f"{x} — {SHORT_NAMES.get(x, str(x))}",
    )

    # Main tab layout (matching design: EDA first, then others)
    tab_eda, tab_abcxyz, tab_forecast, tab_inventory, tab_ts, tab_summary = st.tabs(
        [
            "📊 EDA",
            "🔤 ABC–XYZ",
            "📈 Forecast",
            "📦 Inventory",
            "📉 Time Series",
            "📋 Summary Dashboard",
        ]
    )

    # -------------------------------------------------------------------------
    # EDA TAB
    # -------------------------------------------------------------------------
    with tab_eda:
        st.subheader("📊 Exploratory Data Analysis (EDA)")
        st.write(
            "Baseline historical exploration for the selected material using real inflow and consumption data."
        )

        df_raw, meta = load_material_data(EXCEL_FILE_EOQ, material_code)
        if df_raw is None:
            st.warning("EOQ data file not found or sheet missing.")
        else:
            df_inv = calculate_real_inventory(df_raw)
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Material Code", meta["material_code"])
            with col2:
                st.metric("Material Name", meta["material_name"])
            with col3:
                st.metric("Total Days", len(df_inv))
            with col4:
                st.metric("Total Consumption", f"{df_inv['Tüketim Miktarı'].sum():,.0f}")

            st.markdown("#### Historical Consumption & Inflow")
            fig, ax = plt.subplots(figsize=(12, 4))
            ax.plot(df_inv["Tarih"], df_inv["Tüketim Miktarı"], label="Consumption", color="#DC2626")
            ax.plot(df_inv["Tarih"], df_inv["Mal Giriş"], label="Inflow", color="#16A34A", alpha=0.7)
            ax.set_title("Daily Inflow vs Consumption")
            ax.set_xlabel("Date")
            ax.set_ylabel("Quantity (units)")
            ax.legend()
            ax.grid(True, alpha=0.3)
            st.pyplot(fig)

            st.markdown("#### Inventory Trajectory")
            fig2, ax2 = plt.subplots(figsize=(12, 4))
            ax2.plot(df_inv["Tarih"], df_inv["Envanter"], color="#1f77b4")
            ax2.axhline(0, color="black", linewidth=1, alpha=0.5)
            ax2.set_title("Real Inventory Level Over Time")
            ax2.set_xlabel("Date")
            ax2.set_ylabel("Inventory (units)")
            ax2.grid(True, alpha=0.3)
            st.pyplot(fig2)

            with st.expander("Show raw data"):
                st.dataframe(df_inv)

    # -------------------------------------------------------------------------
    # ABC–XYZ TAB (placeholder structure)
    # -------------------------------------------------------------------------
    with tab_abcxyz:
        st.subheader("🔤 ABC–XYZ Classification")
        st.write(
            "This section is reserved for ABC–XYZ analysis of materials. "
            "You can plug in your classification logic here."
        )
        st.info("Design placeholder: keep layout consistent with other tabs.")

    # -------------------------------------------------------------------------
    # FORECAST TAB
    # -------------------------------------------------------------------------
    with tab_forecast:
        st.subheader("📈 Forecast – Motorin Daily Demand (SARIMA + XGBoost + CatBoost)")
        st.write(
            "Upload a daily Motorin consumption file (columns: 'Tarih', 'Tüketim'). "
            "The system will run rolling one-step forecasts and compare three models."
        )

        uploaded_forecast = st.file_uploader(
            "Upload Motorin Consumption Excel File",
            type=["xlsx", "xls"],
            key="forecast_file",
        )

        if uploaded_forecast is not None:
            try:
                frame = load_daily_from_bytes(uploaded_forecast.read())
                train, test = split_train_test(frame)

                y_train = pd.Series(
                    train[VALUE_COL].values,
                    index=pd.RangeIndex(1, len(train) + 1),
                    name=VALUE_COL,
                )

                if AUTO_SELECT:
                    order, sorder = select_order_aicc(y_train)
                else:
                    order, sorder = DEFAULT_ORDER, DEFAULT_SORDER

                sarima_pred, lo, hi = rolling_one_step(
                    y_train, test[VALUE_COL].values.astype(float), order, sorder
                )

                X_tr, y_tr, feature_cols = build_training_matrix(train)
                xgb_model = fit_xgboost(X_tr, y_tr)
                cat_model = fit_catboost(X_tr, y_tr)

                xgb_pred = rolling_one_step_ml(xgb_model, train, test, feature_cols)
                cat_pred = rolling_one_step_ml(cat_model, train, test, feature_cols)

                comp = build_comparison(test, sarima_pred, lo, hi, xgb_pred, cat_pred)
                stats_dict = {
                    "SARIMA": error_stats(comp["Actual"], comp["SARIMA"]),
                    "XGBoost": error_stats(comp["Actual"], comp["XGBoost"]),
                    "CatBoost": error_stats(comp["Actual"], comp["CatBoost"]),
                }

                col1, col2, col3 = st.columns(3)
                with col1:
                    mae_s, rmse_s, mape_s = stats_dict["SARIMA"]
                    st.metric("SARIMA MAE", f"{mae_s:,.0f}")
                    st.metric("SARIMA RMSE", f"{rmse_s:,.0f}")
                with col2:
                    mae_x, rmse_x, mape_x = stats_dict["XGBoost"]
                    st.metric("XGBoost MAE", f"{mae_x:,.0f}")
                    st.metric("XGBoost RMSE", f"{rmse_x:,.0f}")
                with col3:
                    mae_c, rmse_c, mape_c = stats_dict["CatBoost"]
                    st.metric("CatBoost MAE", f"{mae_c:,.0f}")
                    st.metric("CatBoost RMSE", f"{rmse_c:,.0f}")

                st.markdown("#### Baseline Historical Time Series Data")
                fig_forecast = plot_results(frame, train, comp, order, sorder, stats_dict)
                st.pyplot(fig_forecast)

                st.markdown("#### Forecast Comparison Table")
                st.dataframe(comp)

                excel_bytes = write_excel_forecast(comp, order, sorder, stats_dict)
                st.download_button(
                    label="Download Forecast Comparison Excel",
                    data=excel_bytes,
                    file_name="forecast_motorin_prediction.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error(f"Forecast pipeline failed: {e}")
        else:
            st.info("Please upload a Motorin consumption Excel file to run the forecast comparison.")

    # -------------------------------------------------------------------------
    # INVENTORY TAB
    # -------------------------------------------------------------------------
    with tab_inventory:
        st.subheader("📦 Inventory Optimization – EOQ & ROP with Real Costs")
        df_raw, meta = load_material_data(EXCEL_FILE_EOQ, material_code)
        if df_raw is None:
            st.warning("EOQ data file not found or sheet missing.")
        else:
            df_inv = calculate_real_inventory(df_raw)
            df_costed, cost_summary = calculate_real_inventory_costs(df_inv, meta)
            stats_d = calculate_statistics(df_costed, meta)

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("EOQ (Q*)", f"{stats_d['eoq']:,.0f} units")
            with col2:
                st.metric("ROP", f"{stats_d['rop']:,.0f} units")
            with col3:
                st.metric("Annual Demand", f"{stats_d['annual_demand']:,.0f} units")
            with col4:
                st.metric("Annual Total Cost", f"{cost_summary['currency']}{cost_summary['annual_cost']:,.0f}")

            col5, col6, col7 = st.columns(3)
            with col5:
                st.metric("Total Holding Cost", f"{cost_summary['currency']}{cost_summary['total_holding_cost']:,.0f}")
            with col6:
                st.metric("Total Ordering Cost", f"{cost_summary['currency']}{cost_summary['total_ordering_cost']:,.0f}")
            with col7:
                st.metric("Total Stockout Cost", f"{cost_summary['currency']}{cost_summary['total_stockout_cost']:,.0f}")

            st.markdown("#### EOQ Saw-Tooth Graph")
            fig_eoq = create_eoq_graph(df_costed, stats_d, cost_summary)
            st.pyplot(fig_eoq)

            st.markdown("#### EOQ Cost Tradeoff Function")
            fig_cost = create_eoq_cost_graph(stats_d, cost_summary)
            st.pyplot(fig_cost)

            excel_inv_bytes = export_inventory_cost_data(df_costed, cost_summary)
            st.download_button(
                label="Download Daily Inventory Cost Data (Excel)",
                data=excel_inv_bytes,
                file_name=f"inventory_costs_{meta['material_code']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # -------------------------------------------------------------------------
    # TIME SERIES TAB
    # -------------------------------------------------------------------------
    with tab_ts:
        st.subheader("📉 Time Series – Baseline Historical Data")
        df_raw, meta = load_material_data(EXCEL_FILE_EOQ, material_code)
        if df_raw is None:
            st.warning("EOQ data file not found or sheet missing.")
        else:
            df_inv = calculate_real_inventory(df_raw)
            fig_ts, ax_ts = plt.subplots(figsize=(14, 5))
            ax_ts.plot(df_inv["Tarih"], df_inv["Tüketim Miktarı"], color="#1f77b4", linewidth=1.0)
            ax_ts.set_title("Baseline Historical Time Series Data", fontsize=11, fontweight="bold")
            ax_ts.set_xlabel("Date")
            ax_ts.set_ylabel("Daily Consumption (units)")
            ax_ts.grid(True, alpha=0.3)
            ax_ts.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
            ax_ts.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
            st.pyplot(fig_ts)

    # -------------------------------------------------------------------------
    # SUMMARY DASHBOARD TAB
    # -------------------------------------------------------------------------
    with tab_summary:
        st.subheader("📋 Executive Summary Control Board")

        all_stats = []
        all_costs = []
        for code in MATERIAL_CODES:
            df_raw_i, meta_i = load_material_data(EXCEL_FILE_EOQ, code)
            if df_raw_i is None:
                continue
            df_inv_i = calculate_real_inventory(df_raw_i)
            df_cost_i, cost_i = calculate_real_inventory_costs(df_inv_i, meta_i)
            stats_i = calculate_statistics(df_cost_i, meta_i)
            all_stats.append(stats_i)
            all_costs.append(cost_i)

        if not all_stats:
            st.warning("No materials could be loaded for summary.")
        else:
            df_summary = create_summary_table(all_stats, all_costs)

            target_row = df_summary[df_summary["material_code"] == material_code]
            if not target_row.empty:
                row = target_row.iloc[0]
                left, right = st.columns(2)

                with left:
                    st.markdown("##### Product Profiler Metadata")
                    st.write(f"**Target Inventory Code:** {int(row['material_code'])}")
                    st.write(f"**Description Tag:** {row['material_name']}")
                    st.write(f"**Calculated Theoretical Annual Demand:** {row['annual_demand']:,.0f} units")
                    st.write(f"**Assigned Procurement Lead Time:** {row['lead_time']} days")

                with right:
                    st.markdown("##### Target Recommendation Engine")
                    st.write(f"**Recommended Batch Size (EOQ Q\*):** {row['eoq']:,.0f} units")
                    st.write(f"**Recommended Trigger Threshold (ROP):** {row['rop']:,.0f} units")
                    st.write(f"**Average Daily Cost:** {row['currency']}{row['avg_daily_cost']:,.0f}")
                    st.write(f"**Annual Cost (Real Data):** {row['currency']}{row['annual_cost']:,.0f}")

            st.markdown("#### Portfolio Summary Table")
            st.dataframe(df_summary.reset_index(drop=True))


if __name__ == "__main__":
    main()
